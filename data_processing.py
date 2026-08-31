import random
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

import numpy as np
import torch
from openpyxl import load_workbook
from torch.nn import functional as F
from torch_geometric.data import Batch, Data


Properties = torch.tensor([
    [1.0, 1.0, 1.0],  # A
    [0.0, 0.0, 1.0],  # C
    [1.0, 0.0, 0.0],  # G
    [0.0, 1.0, 0.0],  # T
])
Distanceproperties = torch.tensor([
    [1.0, 1.0, 1.0],  # A
    [0.0, 1.0, 1.0],  # C
    [1.0, 0.0, 0.0],  # G
    [0.0, 0.0, 1.0],  # T
])


@dataclass(frozen=True)
class TaskData:
    name: str
    sequences: np.ndarray
    labels: np.ndarray
    train: np.ndarray
    test: np.ndarray


def read_seq_label(filename):
    sequences, labels = [], []
    workbook = load_workbook(filename, read_only=True, data_only=True)
    if len(workbook.worksheets) < 2:
        raise ValueError(f"Expected two worksheets in {filename}")
    for sheet in workbook.worksheets[:2]:
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] is not None:
                sequences.append(str(row[0]).strip().upper())
                labels.append(int(row[1]))
    workbook.close()
    return sequences, np.asarray(labels, dtype=np.int64)


def encode_sequences(sequences):
    if not sequences or any(len(sequence) != 41 for sequence in sequences):
        raise ValueError("All sequences must contain 41 nucleotides")
    mapping = np.full(256, 255, dtype=np.uint8)
    mapping[[ord(base) for base in "ACGT"]] = np.arange(4, dtype=np.uint8)
    encoded = mapping[np.frombuffer("".join(sequences).encode("ascii"), dtype=np.uint8)]
    encoded = encoded.reshape(-1, 41)
    if np.any(encoded == 255):
        raise ValueError("Sequences must contain only A, C, G and T")
    return encoded


def split_indices(count, seed=2):
    indices = list(range(count))
    rng = random.Random(seed)
    for index in range(count - 1, 0, -1):
        other = int(rng.random() * (index + 1))
        indices[index], indices[other] = indices[other], indices[index]
    train_end = int(0.8 * count)
    validation_end = train_end + int(0.1 * count)
    return (np.asarray(indices[:train_end]),
            np.asarray(indices[train_end:validation_end]),
            np.asarray(indices[validation_end:]))


def load_task_data(filename, name=None):
    sequences, labels = read_seq_label(filename)
    encoded = encode_sequences(sequences)
    train, validation, test = split_indices(len(labels))
    return TaskData(name or Path(filename).stem, encoded, labels,
                    np.concatenate((train, validation)), test)


@lru_cache(maxsize=4)
def _multiscale_edges(length):
    edges = set()
    for position in range(length):
        edges.add((position, position))
        for distance in (1, 2, 4, 8):
            other = position + distance
            if other < length:
                edges.add((position, other))
                edges.add((other, position))
    center = length // 2
    for position in range(length):
        edges.add((center, position))
        edges.add((position, center))
    return torch.tensor(sorted(edges), dtype=torch.long).t().contiguous()


def _local_pair_edges(sequence):
    length = len(sequence)
    edges = {(position, position) for position in range(length)}
    for position in range(length - 1):
        edges.add((position, position + 1))
        edges.add((position + 1, position))
    for first in range(length):
        for second in range(first + 1, min(first + 16, length)):
            pair = (int(sequence[first]), int(sequence[second]))
            if pair in ((0, 3), (3, 0), (1, 2), (2, 1)):
                edges.add((first, second))
                edges.add((second, first))
    return torch.tensor(sorted(edges), dtype=torch.long).t().contiguous()


@lru_cache(maxsize=4)
def _distance10_edges(length):
    edges = []
    for start in range(length - 2):
        for first in range(start, start + 3):
            for second in range(first + 1, start + 3):
                edges.append((first, second))
                edges.append((second, first))
    for first in range(length):
        for second in range(first + 1, min(first + 11, length)):
            edges.append((first, second))
            edges.append((second, first))
    return torch.tensor(edges, dtype=torch.long).t().contiguous()


def build_dna_graph_batch(sequences, graph_mode="multiscale"):
    encoded = np.asarray(sequences)
    if encoded.ndim == 3:
        encoded = encoded.argmax(axis=2)
    if graph_mode not in ("multiscale", "local_pair", "distance10"):
        raise ValueError(
            "Graphmode must be 'multiscale', 'local_pair' or 'distance10'")

    if graph_mode == "multiscale":
        shared_edges = _multiscale_edges(encoded.shape[1])
    elif graph_mode == "distance10":
        shared_edges = _distance10_edges(encoded.shape[1])
    else:
        shared_edges = None
    graphs = []
    for sequence in encoded:
        indices = torch.as_tensor(sequence, dtype=torch.long)
        edges = shared_edges if shared_edges is not None else _local_pair_edges(sequence)
        features = Distanceproperties if graph_mode == "distance10" else Properties
        graphs.append(Data(x=features[indices], edge_index=edges))
    return graphs


def prepare_batch(sequences, graph_mode, device):
    encoded = torch.as_tensor(sequences, dtype=torch.long)
    one_hot = F.one_hot(encoded, num_classes=4).float().to(device)
    graphs = Batch.from_data_list(build_dna_graph_batch(sequences, graph_mode)).to(device)
    return one_hot, graphs


def load_graph_data(filename, graph_mode="multiscale"):
    task = load_task_data(filename)
    train, validation, test = split_indices(len(task.labels))

    def subset(indices):
        sequences = F.one_hot(
            torch.as_tensor(task.sequences[indices], dtype=torch.long), num_classes=4
        ).float().numpy()
        labels = torch.as_tensor(task.labels[indices], dtype=torch.long)
        graphs = build_dna_graph_batch(task.sequences[indices], graph_mode)
        return sequences, labels, graphs

    return (*subset(train), *subset(validation), *subset(test))
